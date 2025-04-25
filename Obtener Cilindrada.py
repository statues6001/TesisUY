import pandas as pd
from openpyxl import load_workbook

# Función para extraer la cilindrada. El codigo no levanta cilindradas inferiores a 1000 cc por lo que las mismas se deben ingresar de forma manual
def extraer_cilindrada_vectorizado(serie_texto):
    return serie_texto.str.extract(r'(\d\.\d)')[0].astype(float).fillna(0) * 1000

# Archivo de entrada y hoja de trabajo. Cambiar el nombre del archivo Excel en caso de ser necesario
archivo_excel = 'Base de datos 2023 - Prueba.xlsx'

# Leer solo la columna 'Modelos' de la hoja '2023'
df = pd.read_excel(archivo_excel, sheet_name='2023', usecols=['Modelos'])

# Extraer cilindrada de manera vectorizada
cilindrada_extraida = extraer_cilindrada_vectorizado(df['Modelos'])

# Cargar el archivo original con openpyxl
workbook = load_workbook(archivo_excel)
hoja = workbook['2023']

# Escribir los datos de la columna 'Cilindrada (cc)' en la siguiente columna vacía
columna_destino = hoja.max_column + 1  # Usar la siguiente columna disponible
hoja.cell(row=1, column=columna_destino, value='Cilindrada (cc)')  # Escribir el encabezado

# Escribir los datos fila por fila
for i, valor in enumerate(cilindrada_extraida, start=2):  # Comienza en la fila 2 (después del encabezado)
    hoja.cell(row=i, column=columna_destino, value=valor)

# Guardar los cambios en el archivo original
workbook.save(archivo_excel)

print(f"Proceso completado. Columna 'Cilindrada (cc)' agregada al archivo '{archivo_excel}'.")
