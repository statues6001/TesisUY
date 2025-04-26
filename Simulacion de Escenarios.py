import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from copy import copy


# ------------------------------------------
# CONFIGURACIONES GLOBALES
# ------------------------------------------
IVA = 0.22  # IVA = 22%
ELASTICIDAD_BASE = -1.6

# Nombres de archivos
archivo_entrada = r"C:\Users\emili\PycharmProjects\TesisUY\Base de datos 2023 - Prueba_output.xlsx"
archivo_salida = "Salida Escenarios.xlsx"

# ------------------------------------------
# LEER DATOS (misma lógica que antes)
# ------------------------------------------
df = pd.read_excel(archivo_entrada)

# ------------------------------------------
# DEFINIR FUNCIONES PARA OBTENER IMESI (%)
# ------------------------------------------
parametros_lineal_esc1 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.00},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00133721, "beta": 0.09761628},

    ("SUV", "N"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("SUV", "HEV"): {"alpha": 0.00133721,  "beta": 0.09761628},
    ("SUV", "MHEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("SUV", "PHEV"): {"alpha": 0.00133721, "beta": 0.09761628},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc2 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.00},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00133721, "beta": 0.09761628},

    ("SUV", "N"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("SUV", "HEV"): {"alpha": 0.00133721,  "beta": 0.09761628},
    ("SUV", "MHEV"): {"alpha": 0.00133721, "beta": 0.09761628},
    ("SUV", "PHEV"): {"alpha": 0.00133721, "beta": 0.09761628},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc3 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.00},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00187204, "beta": 0.0177261},

    ("SUV", "N"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("SUV", "HEV"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("SUV", "MHEV"): {"alpha": 0.00187204, "beta": 0.0177261},
    ("SUV", "PHEV"): {"alpha": 0.00187204, "beta": 0.0177261},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc4 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.00},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00184177, "beta": -0.01533079},

    ("SUV", "N"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("SUV", "HEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "MHEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "PHEV"): {"alpha": 0.00184177, "beta": -0.01533079},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc5 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.00},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00184177, "beta": -0.01533079},

    ("SUV", "N"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("SUV", "HEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "MHEV"): {"alpha": 0.00184177, "beta": -0.01533079},
    ("SUV", "PHEV"): {"alpha": 0.00184177, "beta": -0.01533079},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.00},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc6 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.0},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00264254, "beta": -0.12199636},

    ("SUV", "N"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("SUV", "HEV"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("SUV", "MHEV"): {"alpha": 0.00264254, "beta": -0.12199636},
    ("SUV", "PHEV"): {"alpha": 0.00264254, "beta": -0.12199636},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc7 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.0},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("SUV", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("SUV", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("UTILITARIO", "N"): {"alpha": 0.0, "beta": 0.06},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("UTILITARIO", "HEV"): {"alpha": 0.0, "beta": 0.0115},
    ("UTILITARIO", "MHEV"): {"alpha": 0.0, "beta": 0.0315},
    ("UTILITARIO", "PHEV"): {"alpha": 0.0, "beta": 0.0115}
}

parametros_lineal_esc8 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.0},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("SUV", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("SUV", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("UTILITARIO", "N"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.0},
    ("UTILITARIO", "HEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "MHEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "PHEV"): {"alpha": 0.00093648, "beta": -0.10656341}
}

parametros_lineal_esc9 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.05},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("SUV", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.05},
    ("SUV", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("UTILITARIO", "N"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.05},
    ("UTILITARIO", "HEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "MHEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "PHEV"): {"alpha": 0.00093648, "beta": -0.10656341}
}

parametros_lineal_esc10 = {
    ("AUTOMOVIL", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "D"): {"alpha": 0.0, "beta": 1.15},
    ("AUTOMOVIL", "BEV"): {"alpha": 0.0,   "beta": 0.1},
    ("AUTOMOVIL", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("AUTOMOVIL", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("SUV", "N"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "D"): {"alpha": 0.0, "beta": 1.15},
    ("SUV", "BEV"): {"alpha": 0.0,    "beta": 0.1},
    ("SUV", "HEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "MHEV"): {"alpha": 0.00344331, "beta": -0.22866192},
    ("SUV", "PHEV"): {"alpha": 0.00344331, "beta": -0.22866192},

    ("UTILITARIO", "N"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "D"): {"alpha": 0.0, "beta": 0.347},
    ("UTILITARIO", "BEV"): {"alpha": 0.0,    "beta": 0.1},
    ("UTILITARIO", "HEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "MHEV"): {"alpha": 0.00093648, "beta": -0.10656341},
    ("UTILITARIO", "PHEV"): {"alpha": 0.00093648, "beta": -0.10656341}
}


def imesi_lineal_porcentaje(row, alpha=0.002, beta=0.10):
    co2 = row["CO2 NEDC (g/km)"]
    pct = alpha * co2 + beta
    if pct < 0:
        pct = 0.0
    return pct

def imesi_lineal_por_categoria(row, parametros_lineal):
    # Normalizamos los valores de texto
    tipo2 = str(row["Tipo 2"]).strip().upper()
    motor = str(row["Tipo de motor"]).strip().upper()
    # Si el tipo de motor es BEV, forzar CO2 a 0
    if motor == "BEV":
        co2 = 0
    else:
        try:
            co2 = float(row["CO2 NEDC (g/km)"])
        except Exception:
            return 0.0  # Si hay error en el CO2, devuelve 0.0
    # Consultar el diccionario; si no existe la clave, se usa un valor por defecto
    params = parametros_lineal.get((tipo2, motor), {"alpha": 0.002, "beta": 0.10})
    alpha = params["alpha"]
    beta = params["beta"]
    pct = alpha * co2 + beta

    # Definir topes mínimos y máximos según la combinación (Tipo 2, Tipo de motor)
    topes = {
        ("AUTOMOVIL", "N"): {"min": 0.05, "max": 1},
        ("AUTOMOVIL", "D"): {"min": 0.0, "max": 1.15},
        ("AUTOMOVIL", "BEV"): {"min": 0.0, "max": 1},
        ("AUTOMOVIL", "HEV"): {"min": 0.05, "max": 1},
        ("AUTOMOVIL", "MHEV"): {"min": 0.05, "max": 1},
        ("AUTOMOVIL", "PHEV"): {"min": 0.05, "max": 1},

        ("SUV", "N"): {"min": 0.05, "max": 1},
        ("SUV", "D"): {"min": 0.05, "max": 1.15},
        ("SUV", "BEV"): {"min": 0.0, "max": 1},
        ("SUV", "HEV"): {"min": 0.05, "max": 1},
        ("SUV", "MHEV"): {"min": 0.05, "max": 1},
        ("SUV", "PHEV"): {"min": 0.05, "max": 1},

        ("UTILITARIO", "N"): {"min": 0.06, "max": 0.18},
        ("UTILITARIO", "D"): {"min": 0.347, "max": 0.347},
        ("UTILITARIO", "BEV"): {"min": 0.0, "max": 1},
        ("UTILITARIO", "HEV"): {"min": 0.06, "max": 0.18},
        ("UTILITARIO", "MHEV"): {"min": 0.06, "max": 0.18},
        ("UTILITARIO", "PHEV"): {"min": 0.06, "max": 0.18},
    }

    # Verificar si existe un tope definido para la combinación (Tipo 2, Tipo de motor)
    if (tipo2, motor) in topes:
        pct = max(pct, topes[(tipo2, motor)]["min"])  # Aplica mínimo
        pct = min(pct, topes[(tipo2, motor)]["max"])  # Aplica máximo

    return max(pct, 0)


parametros_escalonados = {
    ("AUTOMOVIL", "N"): [(99, 0.23), (133.4, 0.276), (167.8, 0.322), (202.2, 0.368), (236.6, 0.414), (271, 0.46), (999, 0.46)],
    ("AUTOMOVIL", "D"): [(99999, 1.15)],
    ("AUTOMOVIL", "BEV"): [(99999, 0.00)],
    ("AUTOMOVIL", "HEV"): [(99999, 0.0345)],
    ("AUTOMOVIL", "MHEV"): [(97, 0.07), (117.8, 0.084), (138.6, 0.098), (159.4, 0.112), (180.2, 0.126), (201, 0.14), (999, 0.14)],
    ("AUTOMOVIL", "PHEV"): [(99999, 0.02)],

    ("SUV", "N"): [(99, 0.23), (133.4, 0.276), (167.8, 0.322), (202.2, 0.368), (236.6, 0.414), (271, 0.46), (999, 0.46)],
    ("SUV", "D"): [(99999, 1.15)],
    ("SUV", "BEV"): [(99999, 0.00)],
    ("SUV", "HEV"): [(99999, 0.0345)],
    ("SUV", "MHEV"): [(97, 0.07), (117.8, 0.084), (138.6, 0.098), (159.4, 0.112), (180.2, 0.126), (201, 0.14), (999, 0.14)],
    ("SUV", "PHEV"): [(99999, 0.02)],

    ("UTILITARIO", "N"): [(99999, 0.06)],
    ("UTILITARIO", "D"): [(99999, 0.347)],
    ("UTILITARIO", "BEV"): [(99999, 0.0)],
    ("UTILITARIO", "HEV"): [(99999, 0.0115)],
    ("UTILITARIO", "MHEV"): [(99999, 0.0315)],
    ("UTILITARIO", "PHEV"): [(99999, 0.0115)]
}




def imesi_escalonado_por_categoria(row):
    # Ya se supone que en el DataFrame los valores están normalizados.
    tipo2 = row["Tipo 2"]
    motor = row["Tipo de motor"]
    try:
        co2 = float(row["CO2 NEDC (g/km)"])
    except Exception:
        return 0.0

    if (tipo2, motor) not in parametros_escalonados:
        print("Clave no encontrada:", (tipo2, motor))
        return 0.0

    for umbral, imesi_pct in parametros_escalonados[(tipo2, motor)]:
        if co2 < umbral:
            return imesi_pct
    return 0.0

# ------------------------------------------
# CONFIGURAR RANGO DE ELASTICIDADES
# ------------------------------------------

elasticidades = [-1.1, -1.6, -2.1]

# ------------------------------------------
# FUNCIÓN PRINCIPAL: CALCULAR IMPACTO
# ------------------------------------------
def calcular_impacto_imesi(df_in, funcion_imesi_por_fila, elasticidad, nombre_escenario="Escenario X",
                           elasticidad_variable=False,
                           pct_elasticidad_baratos=0.5,
                           pct_elasticidad_baratos_bajada=1.5):
    df_calc = df_in.copy()
    df_calc["Nuevo IMESI (%)"] = df_calc.apply(funcion_imesi_por_fila, axis=1)
    df_calc["Precio sin Impuestos"] = df_calc["Precio después de tasas"]
    df_calc["Nuevo Monto IMESI (USD)"] = df_calc["Precio sin Impuestos"] * df_calc["Nuevo IMESI (%)"]
    df_calc["Nuevo Precio Final"] = df_calc["Precio sin Impuestos"] * (1 + df_calc["Nuevo IMESI (%)"]) * (1 + IVA)
    df_calc["Precio Final Anterior"] = df_calc["Precio Diciembre 2023 USD"]
    df_calc["Var Precio (%)"] = ((df_calc["Nuevo Precio Final"] - df_calc["Precio Final Anterior"]) / df_calc[
        "Precio Final Anterior"]).fillna(0)

    if elasticidad_variable:
        df_calc["% ranking precio"] = df_calc.groupby("Tipo 2")["Nuevo Precio Final"].rank(pct=True, ascending=True)
        df_calc["Segmento 25 % mas barato"] = df_calc["% ranking precio"] <= 0.25

        def elasticidad_variable_fn(row):
            if row["Segmento 25 % mas barato"]:
                if row["Var Precio (%)"] > 0:
                    return elasticidad * pct_elasticidad_baratos  # Menos negativa
                elif row["Var Precio (%)"] < 0:
                    return elasticidad * pct_elasticidad_baratos_bajada  # Más negativa
                else:
                    return elasticidad
            else:
                return elasticidad

        df_calc["Elasticidad aplicada"] = df_calc.apply(elasticidad_variable_fn, axis=1)
    else:
        # Si no se usa elasticidad variable, asignar la elasticidad original a la columna
        df_calc["Elasticidad aplicada"] = elasticidad

    df_calc["Ventas Escenario"] = df_calc["Procesados"] * (1 + df_calc["Elasticidad aplicada"] * df_calc["Var Precio (%)"])
    df_calc.loc[df_calc["Ventas Escenario"] < 0, "Ventas Escenario"] = 0

    df_calc["Var Ventas (%)"] = ((df_calc["Ventas Escenario"] - df_calc["Procesados"]) / df_calc[
        "Procesados"]).fillna(0)
    df_calc["Recaudación IMESI Escenario"] = df_calc["Nuevo Monto IMESI (USD)"] * df_calc["Ventas Escenario"]
    df_calc["Recaudación IMESI Año Base"] = df_calc["Monto IMESI"] * df_calc["Procesados"]

    resumen = {
        "Escenario": nombre_escenario,
        "Elasticidad": elasticidad,
        "Ventas Año Base (Sin BEV)": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Procesados"].sum(),
        "Ventas BEV Año Base": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Procesados"].sum(),
        "Ventas Escenario (Sin BEV)": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Ventas Escenario"].sum(),
        "Ventas BEV Escenario": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Ventas Escenario"].sum(),
        "Diferencia Ventas (Sin BEV)": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Ventas Escenario"].sum() - df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Procesados"].sum(),
        "Recaudación IMESI Año Base (Sin BEV) (USD)": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Recaudación IMESI Año Base"].sum(),
        "Recaudación IMESI Año Base BEV (USD)": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Recaudación IMESI Año Base"].sum(),
        "Recaudación IMESI Escenario (Sin BEV) (USD)": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Recaudación IMESI Escenario"].sum(),
        "Recaudación IMESI Escenario BEV (USD)": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Recaudación IMESI Escenario"].sum(),
        "Diferencia Recaudación IMESI (USD)": df_calc["Recaudación IMESI Escenario"].sum() - df_calc["Recaudación IMESI Año Base"].sum(),
        "Recaudación IMESI Año Base (Sin BEV) / Unidad": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Recaudación IMESI Año Base"].sum()/df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Procesados"].sum(),
        "Recaudación IMESI Año Base BEV / Unidad": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Recaudación IMESI Año Base"].sum()/df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Procesados"].sum(),
        "Recaudación IMESI Escenario (Sin BEV) / Unidad": df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Recaudación IMESI Escenario"].sum()/df_calc.loc[df_calc["Tipo de motor"] != "BEV", "Ventas Escenario"].sum(),
        "Recaudación IMESI Escenario BEV / Unidad": df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Recaudación IMESI Escenario"].sum()/df_calc.loc[df_calc["Tipo de motor"] == "BEV", "Ventas Escenario"].sum()
    }
    df_calc["Escenario"] = nombre_escenario
    return df_calc, resumen


# ------------------------------------------
# DEFINIR ESCENARIOS (cada escenario será una pestaña nueva)
# ------------------------------------------
escenarios = []

# Carga Escenario Escalonado 1. Elasticidad fija = -1.6
escenarios.append({
    "nombre": "Escalonado Escenario 1",
    "func_imesi": imesi_escalonado_por_categoria,
    "elasticidad": ELASTICIDAD_BASE
})

# Carga Escenario Lineal 1
for elasticidad in elasticidades:
    escenarios.append({
        "nombre": f"Lineal Escenario 1 (E={elasticidad})",
        "func_imesi": lambda row, p=parametros_lineal_esc1: imesi_lineal_por_categoria(row, p),
        "elasticidad": elasticidad
    })

# Carga Escenario Lineal 2 (elasticidad variable)
for elasticidad in elasticidades:
    escenarios.append({
        "nombre": f"Lineal Escenario 2 (E={elasticidad})",
        "func_imesi": lambda row, p=parametros_lineal_esc2: imesi_lineal_por_categoria(row, p),
        "elasticidad": elasticidad,
        "elasticidad_variable": True,
        "pct_elasticidad_baratos": 0.5,
        "pct_elasticidad_baratos_bajada": 1.5
    })

# Carga Escenario Lineal 3
for elasticidad in elasticidades:
    escenarios.append({
        "nombre": f"Lineal Escenario 3 (E={elasticidad})",
        "func_imesi": lambda row, p=parametros_lineal_esc3: imesi_lineal_por_categoria(row, p),
        "elasticidad": elasticidad
    })

# Carga Escenario Lineal 4
for elasticidad in elasticidades:
    escenarios.append({
        "nombre": f"Lineal Escenario 4 (E={elasticidad})",
        "func_imesi": lambda row, p=parametros_lineal_esc4: imesi_lineal_por_categoria(row, p),
        "elasticidad": elasticidad
    })

# Carga Escenario Lineal 5 (elasticidad variable)
for elasticidad in elasticidades:
    escenarios.append({
        "nombre": f"Lineal Escenario 5 (E={elasticidad})",
        "func_imesi": lambda row, p=parametros_lineal_esc5: imesi_lineal_por_categoria(row, p),
        "elasticidad": elasticidad,
        "elasticidad_variable": True,
        "pct_elasticidad_baratos": 0.5,
        "pct_elasticidad_baratos_bajada": 1.5
    })

# Carga Escenario Lineal 6
for elasticidad in elasticidades:
    escenarios.append({
            "nombre": f"Lineal Escenario 6 (E={elasticidad})",
            "func_imesi": lambda row, p=parametros_lineal_esc6: imesi_lineal_por_categoria(row, p),
            "elasticidad": elasticidad
    })

# Carga Escenario Lineal 7
for elasticidad in elasticidades:
    escenarios.append({
            "nombre": f"Lineal Escenario 7 (E={elasticidad})",
            "func_imesi": lambda row, p=parametros_lineal_esc7: imesi_lineal_por_categoria(row, p),
            "elasticidad": elasticidad
    })

# Carga Escenario Lineal 8
for elasticidad in elasticidades:
    escenarios.append({
            "nombre": f"Lineal Escenario 8 (E={elasticidad})",
            "func_imesi": lambda row, p=parametros_lineal_esc8: imesi_lineal_por_categoria(row, p),
            "elasticidad": elasticidad
    })

# Carga Escenario Lineal 9
for elasticidad in elasticidades:
    escenarios.append({
            "nombre": f"Lineal Escenario 9 (E={elasticidad})",
            "func_imesi": lambda row, p=parametros_lineal_esc9: imesi_lineal_por_categoria(row, p),
            "elasticidad": elasticidad
    })

# Carga Escenario Lineal 10
for elasticidad in elasticidades:
    escenarios.append({
            "nombre": f"Lineal Escenario 10 (E={elasticidad})",
            "func_imesi": lambda row, p=parametros_lineal_esc10: imesi_lineal_por_categoria(row, p),
            "elasticidad": elasticidad
    })


# Ordenar escenarios por nombre
#escenarios.sort(key=lambda x: (x["nombre"]))

# Generar resultados para cada escenario
resultados_resumen = []
resultados_detallados = {}  # Diccionario: clave = nombre del escenario, valor = DataFrame

for esc in escenarios:
    df_res, resumen = calcular_impacto_imesi(
        df,
        esc["func_imesi"],
        esc["elasticidad"],
        esc["nombre"],
        elasticidad_variable=esc.get("elasticidad_variable", False),
        pct_elasticidad_baratos=esc.get("pct_elasticidad_baratos", 0.5),
        pct_elasticidad_baratos_bajada=esc.get("pct_elasticidad_baratos_bajada", 1.5)
    )

    # Si el escenario es "Lineal por Categoría", actualizamos la columna "Escenario"
    if esc["nombre"] == "Lineal por Categoría":
        def get_alfa_beta(row):
            # Normalizamos los valores para buscar en el diccionario
            tipo2 = str(row["Tipo 2"]).strip().upper()
            motor = str(row["Tipo de motor"]).strip().upper()
            params = parametros_lineal_esc1.get((tipo2, motor), {"alpha": 0.002, "beta": 0.10})
            return f"alfa: {params['alpha']}, beta: {params['beta']}"


        df_res["Escenario"] = df_res.apply(get_alfa_beta, axis=1)
    else:
        df_res["Escenario"] = esc["nombre"]

    resultados_detallados[esc["nombre"]] = df_res.copy()
    resultados_resumen.append(resumen)

df_resumen = pd.DataFrame(resultados_resumen)

# ------------------------------------------
# USAR OPENPYXL PARA CREAR UNA PESTAÑA NUEVA POR ESCENARIO
# ------------------------------------------
wb = load_workbook(archivo_entrada)
ws_original = wb["2023"]

#Definición de estilos para fila 1 (encabezado)

ref_header = ws_original.cell(row=1, column=1)
header_font = ref_header.font
header_border = ref_header.border
header_fill = ref_header.fill
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Para el resto de los datos, definir una alineación centrada y un borde fino
center_alignment = Alignment(horizontal="center", vertical="center")
thin_side = Side(style="thin", color="000000")
data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

# Si ya existen hojas con los nombres que vamos a usar, eliminarlas
for esc in escenarios:
    sheet_name = f"{esc['nombre']}"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
if "Resumen Escenarios" in wb.sheetnames:
    del wb["Resumen Escenarios"]

# Para cada escenario, copiar la hoja original y agregar las columnas nuevas
for esc in escenarios:
    sheet_name = f"{esc['nombre']}"
    # Copiar la hoja original (con fórmulas y formato)
    ws_new = wb.copy_worksheet(ws_original)
    ws_new.title = sheet_name

    # Obtener el DataFrame de resultados para este escenario
    df_det = resultados_detallados[esc["nombre"]]

    # Determinar cuántas columnas tenía la hoja original
    last_col = ws_original.max_column
    new_columns = list(df_det.columns)[last_col:]

    # Escribir encabezados para las nuevas columnas a partir de la columna last_col+1
    for i, col_name in enumerate(new_columns, start=last_col + 1):
        new_header = ws_new.cell(row=1, column=i, value=col_name)
        new_header.font = copy(header_font)
        new_header.border = copy(header_border)
        new_header.fill = copy(header_fill)
        new_header.alignment = copy(header_alignment)

    # Escribir los datos de las nuevas columnas para cada fila
    for row_idx, row in df_det.iterrows():
        # Asumimos que la fila 0 de df_det corresponde a la fila 2 en Excel (encabezados en la fila 1)
        for i, col_name in enumerate(new_columns, start=last_col + 1):
            valor = row[col_name]
            cell = ws_new.cell(row=row_idx + 2, column=i, value=valor if pd.notna(valor) else None)
            cell.alignment = center_alignment
            cell.border = data_border

            if col_name in ["Nuevo IMESI (%)", "Var Precio (%)", "Var Ventas (%)"]:
                cell.number_format = '0.00%'


# ---------------------------
# ANALISIS ENERGETICO
# ---------------------------
# Se creará un DataFrame maestro que contenga, para cada escenario,
# el consumo ponderado (viejo y nuevo) y las ventas (viejas y nuevas) agrupados por "Tipo 2" y "Tipo de motor".

df_master = pd.DataFrame()

def safe_divide(numerador, denominador):
    return numerador / denominador if denominador != 0 else 0

for esc in escenarios:
    # Obtener el DataFrame del escenario (para las Ventas Escenario)
    df_new = resultados_detallados[esc["nombre"]]

    agg_old = df.groupby(["Tipo 2", "Tipo de motor"])[["Procesados", "Consumo (L/100 km)"]].apply(
        lambda g: pd.Series({
            "Ventas Año Base": g["Procesados"].sum(),
            "Consumo Año Base (L/100 km)": safe_divide((g["Consumo (L/100 km)"] * g["Procesados"]).sum(),
                                                       g["Procesados"].sum())
        })
    ).reset_index()

    agg_new = df_new.groupby(["Tipo 2", "Tipo de motor"])[["Ventas Escenario", "Consumo (L/100 km)"]].apply(
        lambda g: pd.Series({
            "Ventas Escenario": g["Ventas Escenario"].sum(),
            "Consumo Escenario (L/100 km)": safe_divide((g["Consumo (L/100 km)"] * g["Ventas Escenario"]).sum(),
                                                        g["Ventas Escenario"].sum())
        })
    ).reset_index()

    # Unir ambos resúmenes
    df_analisis = pd.merge(agg_old, agg_new, on=["Tipo 2", "Tipo de motor"], how="outer")
    # Agregar la columna que identifica el escenario
    df_analisis["Escenario"] = esc["nombre"]
    # Reorganizar columnas para mayor claridad
    df_analisis = df_analisis[["Escenario", "Tipo 2", "Tipo de motor", "Consumo Año Base (L/100 km)", "Ventas Año Base",
                               "Consumo Escenario (L/100 km)", "Ventas Escenario"]]

    # Acumular el resultado en el DataFrame maestro
    df_master = pd.concat([df_master, df_analisis], ignore_index=True)

# ---------------------------
# Agregar columnas de kilómetros recorridos y litros consumidos
# Definir los Recorrido anual (km) anuales según "Tipo 2"
km_por_tipo = {
    "AUTOMOVIL": 10252,
    "SUV": 10252,
    "UTILITARIO": 31200
}
# Asignar los Recorrido anual (km) según el valor de "Tipo 2"
df_master["Recorrido anual (km)"] = df_master["Tipo 2"].map(km_por_tipo)

# Agregar manejo especial únicamente para vehículos eléctricos puros (BEV)
df_master["Consumo eléctrico Año Base (kWh/km)"] = None
df_master["Consumo eléctrico Escenario (kWh/km)"] = None
df_master["Variación consumo eléctrico (kWh)"] = None
df_master["Consumo eléctrico Escenario (kWh)"] = None
df_master["Consumo eléctrico Escenario (BEV) (tep)"] = None

# Procesar cada fila
for idx, row in df_master.iterrows():
    tipo_motor = row["Tipo de motor"]
    tipo2 = row["Tipo 2"]

    if tipo_motor == "BEV":
        # Filtramos solo los vehículos eléctricos de la misma categoría (Automóvil, SUV o Utilitario)
        df_electricos = df[(df["Tipo de motor"] == tipo_motor) & (df["Tipo 2"] == tipo2)]

        # Calcular Consumo eléctrico Año Base
        rendimiento_viejo = (df_electricos["Rendimiento eléctrico (km/kWh)"] * df_electricos["Procesados"]).sum() / \
                            df_electricos["Procesados"].sum()
        consumo_viejo = 1 / rendimiento_viejo  # (kWh/km)

        # Guardar resultado de consumo eléctrico año base (kWh/km)
        df_master.at[idx, "Consumo eléctrico Año Base (kWh/km)"] = consumo_viejo

        # Calcular y guardar Consumo eléctrico Año Base total (kWh)
        consumo_electrico_anio_base_kWh = row["Ventas Año Base"] * consumo_viejo * row["Recorrido anual (km)"]
        df_master.at[idx, "Consumo eléctrico Año Base (kWh)"] = consumo_electrico_anio_base_kWh

        # Convertir a tep (0.086 tep/MWh)
        consumo_electrico_anio_base_tep = (consumo_electrico_anio_base_kWh / 1000) * 0.086
        df_master.at[idx, "Consumo eléctrico Año Base (BEV) (tep)"] = consumo_electrico_anio_base_tep

        # Calcular Consumo eléctrico Escenario nuevo (usando Ventas Escenario)
        df_electricos_nuevo = resultados_detallados[row["Escenario"]]
        df_electricos_nuevo = df_electricos_nuevo[(df_electricos_nuevo["Tipo de motor"] == tipo_motor) &
                                                  (df_electricos_nuevo["Tipo 2"] == tipo2)]
        rendimiento_nuevo = (df_electricos_nuevo["Rendimiento eléctrico (km/kWh)"] *
                             df_electricos_nuevo["Ventas Escenario"]).sum() / df_electricos_nuevo[
                                "Ventas Escenario"].sum()
        consumo_nuevo = 1 / rendimiento_nuevo  # (kWh/km)

        # Guardar resultados del consumo eléctrico Escenario (kWh/km)
        df_master.at[idx, "Consumo eléctrico Escenario (kWh/km)"] = consumo_nuevo

        # Verificar que el consumo eléctrico del escenario (kWh/km) tenga valor
        consumo_kWh_km = df_master.at[idx, "Consumo eléctrico Escenario (kWh/km)"]
        if pd.isna(consumo_kWh_km):
            consumo_kWh_km = 0

        # Calcular el consumo eléctrico total en kWh para el escenario
        consumo_electrico_kWh = row["Ventas Escenario"] * consumo_kWh_km * row["Recorrido anual (km)"]
        df_master.at[idx, "Consumo eléctrico Escenario (kWh)"] = consumo_electrico_kWh

        # Convertir de kWh a tep usando la relación: 0.086 tep/MWh
        consumo_electrico_tep = (consumo_electrico_kWh / 1000) * 0.086
        df_master.at[idx, "Consumo eléctrico Escenario (BEV) (tep)"] = consumo_electrico_tep

        # Calcular Variación consumo eléctrico (kWh)
        variacion_kwh = consumo_electrico_kWh - consumo_electrico_anio_base_kWh
        df_master.at[idx, "Variación consumo eléctrico (kWh)"] = round(variacion_kwh, 2)

# Calcular litros consumidos para el escenario viejo y el nuevo
df_master["Consumo Año Base (L)"] = df_master["Ventas Año Base"] * (
    df_master["Consumo Año Base (L/100 km)"] * df_master["Recorrido anual (km)"] / 100)

# Agregar la columna "tep consumidos Año Base"
df_master["Consumo Año Base (sin BEV) (tep)"] = df_master.apply(
    lambda row: (row["Consumo Año Base (L)"] / 1000) * (0.8551 if row["Tipo de motor"] == "D" else 0.7774),
    axis=1
)

df_master["Consumo Escenario (L)"] = df_master["Ventas Escenario"] * (
        df_master["Consumo Escenario (L/100 km)"] * df_master["Recorrido anual (km)"] / 100)

# Agregar la columna "tep consumidos Escenario"
df_master["Consumo Escenario (sin BEV) (tep)"] = df_master.apply(
    lambda row: (row["Consumo Escenario (L)"] / 1000) * (0.8551 if row["Tipo de motor"] == "D" else 0.7774),
    axis=1
)

df_master["Variación consumo anual (L)"] = df_master["Consumo Escenario (L)"] - df_master["Consumo Año Base (L)"]
df_master["Variación consumo anual (L)"] = df_master["Variación consumo anual (L)"].round(4)

df_master["Variación consumo anual (m3)"] = df_master["Variación consumo anual (L)"]/1000

def calcular_emisiones(row):
    if row["Tipo de motor"] == "BEV":
        # Para BEV se usa el consumo eléctrico
        emis_bev_base = (row["Consumo eléctrico Año Base (kWh)"] / 1e6) * 56
        emis_bev_escenario = (row["Consumo eléctrico Escenario (kWh)"] / 1e6) * 56
        emis_sin_bev_base = 0
        emis_sin_bev_escenario = 0
    else:
        emis_bev_base = 0
        emis_bev_escenario = 0
        # Factor según el tipo de combustible: Gasoil para "D", Gasolina para el resto
        factor = 2684.4018 if row["Tipo de motor"] == "D" else 2336.8676
        emis_sin_bev_base = (row["Consumo Año Base (L)"] * factor) / 1e6
        emis_sin_bev_escenario = (row["Consumo Escenario (L)"] * factor) / 1e6
    emis_total_base = emis_bev_base + emis_sin_bev_base
    emis_total_escenario = emis_bev_escenario + emis_sin_bev_escenario
    variacion = emis_total_escenario - emis_total_base
    return pd.Series({
        "Emisiones BEV CO2 Año Base (ton)": emis_bev_base,
        "Emisiones BEV CO2 Escenario (ton)": emis_bev_escenario,
        "Emisiones Sin BEV CO2 Año Base (ton)": emis_sin_bev_base,
        "Emisiones Sin BEV CO2 Escenario (ton)": emis_sin_bev_escenario,
        "Emisiones CO2 Año Base (ton)": emis_total_base,
        "Emisiones CO2 Escenario (ton)": emis_total_escenario,
        "Variación CO2 (ton)": variacion
    })

emisiones = df_master.apply(calcular_emisiones, axis=1)
df_master = pd.concat([df_master, emisiones], axis=1)

# Calcular la variación de emisiones de CO2 (ton)
df_master["Variación CO2 (ton)"] = df_master["Emisiones CO2 Escenario (ton)"] - df_master["Emisiones CO2 Año Base (ton)"]
df_master["Variación CO2 (ton)"] = df_master["Variación CO2 (ton)"].round(6)

# Reorganizar columnas para una presentación clara
df_master = df_master[["Escenario", "Tipo 2", "Tipo de motor", "Recorrido anual (km)",
    "Ventas Año Base", "Ventas Escenario", "Consumo Año Base (L/100 km)", "Consumo Año Base (L)", "Consumo Año Base (sin BEV) (tep)",
    "Consumo Escenario (L/100 km)", "Consumo Escenario (L)", "Consumo Escenario (sin BEV) (tep)", "Variación consumo anual (L)",
    "Variación consumo anual (m3)", "Consumo eléctrico Año Base (kWh/km)", "Consumo eléctrico Año Base (kWh)", "Consumo eléctrico Año Base (BEV) (tep)",
    "Consumo eléctrico Escenario (kWh/km)", "Consumo eléctrico Escenario (kWh)", "Consumo eléctrico Escenario (BEV) (tep)",
    "Variación consumo eléctrico (kWh)", "Emisiones Sin BEV CO2 Año Base (ton)", "Emisiones Sin BEV CO2 Escenario (ton)",
    "Emisiones BEV CO2 Año Base (ton)", "Emisiones BEV CO2 Escenario (ton)",
    "Emisiones CO2 Año Base (ton)", "Emisiones CO2 Escenario (ton)", "Variación CO2 (ton)"]]


# Calcular columna "Consumo Año Base (sin BEV) (tep)/Unidad"
tep_base_sin_bev_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros que no sean BEV para Año Base
    df_temp_base = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] != "BEV")]
    total_tep_base = df_temp_base["Consumo Año Base (sin BEV) (tep)"].sum()
    ventas_base_sin_bev = df_temp_base["Ventas Año Base"].sum()
    tep_base_sin_bev_por_escenario[esc] = total_tep_base / ventas_base_sin_bev if ventas_base_sin_bev != 0 else 0

# Actualizar el DataFrame resumen

df_resumen["Consumo Año Base (sin BEV) (tep)/Unidad"] = df_resumen["Escenario"].map(tep_base_sin_bev_por_escenario)

# Calcular la nueva columna "Consumo Escenario (sin BEV) (tep)/Unidad" para cada escenario
tep_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo los registros que no sean BEV
    df_temp = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] != "BEV")]
    total_tep = df_temp["Consumo Escenario (sin BEV) (tep)"].sum()
    # Se usan las Ventas Escenario para calcular el valor por unidad (equivalente a "Ventas Escenario (Sin BEV)")
    ventas_escenario_sin_bev = df_temp["Ventas Escenario"].sum()
    tep_por_escenario[esc] = total_tep / ventas_escenario_sin_bev if ventas_escenario_sin_bev != 0 else 0

# Actualizar el DataFrame de resumen (df_resumen) agregando la nueva columna, mapeando por el nombre del escenario
df_resumen["Consumo Escenario (sin BEV) (tep)/Unidad"] = df_resumen["Escenario"].map(tep_por_escenario)

# Calcular columna "Consumo eléctrico Año Base BEV (tep)/Unidad"
tep_base_bev_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros BEV para Año Base
    df_temp_base_bev = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] == "BEV")]
    total_tep_base_bev = df_temp_base_bev["Consumo eléctrico Año Base (BEV) (tep)"].sum()
    ventas_base_bev = df_temp_base_bev["Ventas Año Base"].sum()
    tep_base_bev_por_escenario[esc] = total_tep_base_bev / ventas_base_bev if ventas_base_bev != 0 else 0

# Actualizar el DataFrame resumen

df_resumen["Consumo eléctrico Año Base BEV (tep)/Unidad"] = df_resumen["Escenario"].map(tep_base_bev_por_escenario)

# Calcular la nueva columna "Consumo eléctrico Escenario BEV (tep)/Unidad"
tep_bev_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo los registros que sean BEV
    df_temp = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] == "BEV")]
    total_tep = df_temp["Consumo eléctrico Escenario (BEV) (tep)"].sum()
    # Se usan las Ventas Escenario para calcular el valor por unidad (para BEV)
    ventas_escenario_bev = df_temp["Ventas Escenario"].sum()
    tep_bev_por_escenario[esc] = total_tep / ventas_escenario_bev if ventas_escenario_bev != 0 else 0

# Actualizar el DataFrame de resumen (df_resumen) agregando la nueva columna, mapeando por el nombre del escenario
df_resumen["Consumo eléctrico Escenario BEV (tep)/Unidad"] = df_resumen["Escenario"].map(tep_bev_por_escenario)

# Emisiones BEV CO2 Año Base (ton)/Unidad
emisiones_bev_base_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros BEV para Año Base
    df_temp_bev = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] == "BEV")]
    total_emisiones_bev_base = df_temp_bev["Emisiones BEV CO2 Año Base (ton)"].sum()
    ventas_base_bev = df_temp_bev["Ventas Año Base"].sum()
    emisiones_bev_base_por_escenario[esc] = total_emisiones_bev_base / ventas_base_bev if ventas_base_bev != 0 else 0

# Actualizar el DataFrame resumen
df_resumen["Emisiones BEV CO2 Año Base (ton)/Unidad"] = df_resumen["Escenario"].map(emisiones_bev_base_por_escenario)

# Emisiones BEV CO2 Escenario (ton)/Unidad
emisiones_bev_escenario_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros BEV para Escenario
    df_temp_bev = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] == "BEV")]
    total_emisiones_bev_escenario = df_temp_bev["Emisiones BEV CO2 Escenario (ton)"].sum()
    ventas_escenario_bev = df_temp_bev["Ventas Escenario"].sum()
    emisiones_bev_escenario_por_escenario[esc] = total_emisiones_bev_escenario / ventas_escenario_bev if ventas_escenario_bev != 0 else 0

df_resumen["Emisiones BEV CO2 Escenario (ton)/Unidad"] = df_resumen["Escenario"].map(emisiones_bev_escenario_por_escenario)

# Emisiones Sin BEV CO2 Año Base (ton)/Unidad
emisiones_sin_bev_base_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros que no sean BEV para Año Base
    df_temp_sin_bev = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] != "BEV")]
    total_emisiones_sin_bev_base = df_temp_sin_bev["Emisiones Sin BEV CO2 Año Base (ton)"].sum()
    ventas_base_sin_bev = df_temp_sin_bev["Ventas Año Base"].sum()
    emisiones_sin_bev_base_por_escenario[esc] = total_emisiones_sin_bev_base / ventas_base_sin_bev if ventas_base_sin_bev != 0 else 0

df_resumen["Emisiones Sin BEV CO2 Año Base (ton)/Unidad"] = df_resumen["Escenario"].map(emisiones_sin_bev_base_por_escenario)

# Emisiones Sin BEV CO2 Escenario (ton)/Unidad
emisiones_sin_bev_escenario_por_escenario = {}
for esc in df_master["Escenario"].unique():
    # Filtrar solo registros que no sean BEV para Escenario
    df_temp_sin_bev = df_master[(df_master["Escenario"] == esc) & (df_master["Tipo de motor"] != "BEV")]
    total_emisiones_sin_bev_escenario = df_temp_sin_bev["Emisiones Sin BEV CO2 Escenario (ton)"].sum()
    ventas_escenario_sin_bev = df_temp_sin_bev["Ventas Escenario"].sum()
    emisiones_sin_bev_escenario_por_escenario[esc] = total_emisiones_sin_bev_escenario / ventas_escenario_sin_bev if ventas_escenario_sin_bev != 0 else 0

df_resumen["Emisiones Sin BEV CO2 Escenario (ton)/Unidad"] = df_resumen["Escenario"].map(emisiones_sin_bev_escenario_por_escenario)

# ------------------------------------------
# CREAR PESTAÑA "Resumen Escenarios". Se crea en este punto ya que hay dos columnas del df_resumen que depende de los
# datos de la pestaña de análisis energético contenidos en df_master
# ------------------------------------------
ws_resumen = wb.create_sheet("Resumen Escenarios")
for col_idx, col_name in enumerate(df_resumen.columns, start=1):
    cell = ws_resumen.cell(row=1, column=col_idx, value=col_name)
    cell.font = copy(header_font)
    cell.border = copy(header_border)
    cell.fill = copy(header_fill)
    cell.alignment = copy(header_alignment)

for row_idx, row in df_resumen.iterrows():
    for col_idx, cell_value in enumerate(row, start=1):
        cell = ws_resumen.cell(row=row_idx + 2, column=col_idx, value=cell_value)
        cell.alignment = center_alignment
        cell.border = data_border

# ------------------------------------------
# CREAR PESTAÑA única para el Análisis Energético
# ------------------------------------------
if "Analisis energético" in wb.sheetnames:
    del wb["Analisis energético"]
ws_energetico = wb.create_sheet("Analisis energético")

# Escribir los encabezados en la nueva pestaña
for col_idx, col_name in enumerate(df_master.columns, start=1):
    cell = ws_energetico.cell(row=1, column=col_idx, value=col_name)
    cell.font = copy(header_font)
    cell.border = copy(header_border)
    cell.fill = copy(header_fill)
    cell.alignment = copy(header_alignment)

# Escribir los datos del resumen concentrado
for row_idx, row in df_master.iterrows():
    for col_idx, value in enumerate(row, start=1):
        cell = ws_energetico.cell(row=row_idx + 2, column=col_idx, value=value if pd.notna(value) else None)
        cell.alignment = center_alignment
        cell.border = data_border

# ---------------------------------------------------------------------------------------
# Reordenar hojas para que "Analisis energético" y "Resumen Escenarios" aparezcan primero
# ---------------------------------------------------------------------------------------
ws_energetico = wb["Analisis energético"]
ws_resumen = wb["Resumen Escenarios"]
resto_hojas = [hoja for hoja in wb.worksheets if hoja not in [ws_energetico, ws_resumen]]
wb._sheets = [ws_energetico, ws_resumen] + resto_hojas
# ------------------------------------------
# GUARDAR ARCHIVO FINAL
# ------------------------------------------
wb.save(archivo_salida)
print(
    f"Proceso completado. Se crearon hojas nuevas para cada escenario y la pestaña 'Resumen Escenarios' en '{archivo_salida}'.")
