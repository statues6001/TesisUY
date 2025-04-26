import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from copy import copy

def procesar_imesi(file_path, sheet_name="2023"):

    # 1) Lectura y limpieza de datos
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    # Se revisa que existan estas columnas en el Excel (ajustar los nombres si difieren)
    # Limpieza: forzar texto y numéricos
    df["Tipo 1"] = df["Tipo 1"].fillna("DESCONOCIDO").astype(str)
    df["Combustible 2"] = df["Combustible 2"].fillna("DESCONOCIDO").astype(str)
    df["Tipo de motor"] = df["Tipo de motor"].fillna("DESCONOCIDO").astype(str)
    df["Cilindrada"] = pd.to_numeric(df["Cilindrada"], errors="coerce").fillna(0.0)
    df["Precio Diciembre 2023 USD"] = pd.to_numeric(
        df["Precio Diciembre 2023 USD"], errors="coerce"
    ).fillna(0.0)

    # IVA fijo en 22 %
    IVA = 0.22

    # 2) Calulos.
    # Se generan arrays para guardar resultados
    monto_imesi_list = []
    imesi_fraction_list = []
    precio_despues_tasas_list = []

    for idx, row in df.iterrows():
        # Extraemos campos
        tipo = row["Tipo 1"].upper().strip()
        cilindrada = row["Cilindrada"]
        combustible = row["Combustible 2"].upper().strip()
        tipo_motor = row["Tipo de motor"].upper().strip()
        precio_mercado = row["Precio Diciembre 2023 USD"]

        # 2.1 Fundión para determinar fracción de IMESI
        imesi_frac = determinar_imesi_fraction(tipo, cilindrada, combustible, tipo_motor)

        # 2.2 Calcular Precio después de tasas
        #     precio_mercado = (precio_despues_tasas * (1 + imesi_frac)) * (1 + IVA)
        # =>  precio_despues_tasas = precio_mercado / ((1 + imesi_frac) * (1 + IVA))
        if (1 + imesi_frac) * (1 + IVA) == 0:
            precio_despues_de_tasas = 0.0
        else:
            precio_despues_de_tasas = precio_mercado / ((1 + imesi_frac) * (1 + IVA))

        # 2.3 Calcular monto IMESI en dólares
        #     Monto IMESI = precio_despues_tasas * imesi_frac
        monto_imesi = precio_despues_de_tasas * imesi_frac

        # Guardamos en listas
        monto_imesi_list.append(monto_imesi)
        imesi_fraction_list.append(imesi_frac)
        precio_despues_tasas_list.append(precio_despues_de_tasas)

    # --------------------------------------------------------
    # 3) Insertar las 3 nuevas "columnas" en el DataFrame
    # --------------------------------------------------------
    df["Monto IMESI"] = monto_imesi_list
    df["IMESI_frac"] = imesi_fraction_list
    df["Precio después de tasas"] = precio_despues_tasas_list

    # --------------------------------------------------------
    # 4) Escribir resultados en un nuevo archivo (para no sobreescribir el original)
    # --------------------------------------------------------
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # Definición de estilos para fila 1 (encabezado)

    ref_header = ws.cell(row=1, column=1)
    header_font = ref_header.font
    header_border = ref_header.border
    header_fill = ref_header.fill
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Para el resto de los datos, definir una alineación centrada y un borde fino
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1) Hallar la última columna usada
    last_used_col = ws.max_column
    start_col = last_used_col + 1

    col_monto_imesi = start_col
    col_imesi_pct = start_col + 1
    col_precio_net = start_col + 2

    # 2) Encabezados en la fila 1
    ws.cell(row=1, column=col_monto_imesi, value="Monto IMESI")
    ws.cell(row=1, column=col_imesi_pct, value="IMESI (%)")
    ws.cell(row=1, column=col_precio_net, value="Precio después de tasas")

    headers = {
        col_monto_imesi: "Monto IMESI",
        col_imesi_pct: "IMESI (%)",
        col_precio_net: "Precio después de tasas"
    }

    for col, header in headers.items():
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = copy(header_font)
        cell.border = copy(header_border)
        cell.fill = copy(header_fill)
        cell.alignment = header_alignment


    # 3) Escribir datos fila a fila
    start_row = 2
    for i, row_df in df.iterrows():
        excel_row = i + start_row

        cell_ime = ws.cell(row=excel_row, column=col_monto_imesi, value=row_df["Monto IMESI"])
        cell_ime.alignment = center_alignment
        cell_ime.border = data_border
        c_pct = ws.cell(row=excel_row, column=col_imesi_pct, value=row_df["IMESI_frac"])
        c_pct.number_format = '0.00%'
        c_pct.alignment = center_alignment
        c_pct.border = data_border
        cell2 = ws.cell(row=excel_row, column=col_precio_net, value=row_df["Precio después de tasas"])
        cell2.alignment = center_alignment
        cell2.border = data_border

    # Guardamos en archivo nuevo, para no pisar el original
    new_file_path = file_path.replace(".xlsx", "_output.xlsx")
    wb.save(new_file_path)
    print(f"Proceso completado. Revisa el archivo: {new_file_path}")


def determinar_imesi_fraction(tipo, cilindrada, combustible, tipo_motor):
    """
    Lógica que retorna la fracción de IMESI (por ejemplo, 0.347 para 34,7%).
    """
    # Por defecto 0
    imesi_str = "0"

    # -----------------------------
    # TIPO = COMERCIAL
    # -----------------------------
    if tipo == "COMERCIAL":
        # Combustible = E -> 0%
        if combustible == "E":
            imesi_str = "0"
        else:
            # a) Cilindrada <= 1600
            if cilindrada <= 1600:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "34,7"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "6,0"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"

            # b) 1600 < cilindrada <= 3500
            elif 1600 < cilindrada <= 3500:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "34,7"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "6,0"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"

            # c) cilindrada > 3500
            elif cilindrada > 3500:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "80,5"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "11,5"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "1,15"
                    elif tipo_motor == "MHEV":
                        imesi_str = "3,15"

            # d) Sin más rangos => 0
            else:
                imesi_str = "0"

    # -----------------------------
    # TIPO = AUTOMOVIL
    # -----------------------------
    elif tipo == "AUTOMOVIL":
        # Combustible = E -> 0%
        if combustible == "E":
            imesi_str = "0"
        else:
            # a) Cilindrada <= 1000
            if cilindrada <= 1000:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    elif tipo_motor == "PHEV":
                        imesi_str = "2"
                    elif tipo_motor == "HEV":
                        imesi_str = "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "7"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "23,00"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "2" if tipo_motor == "PHEV" else "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "7"

            # b) 1000 < cilindrada <= 1500
            elif 1000 < cilindrada <= 1500:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "2" if tipo_motor == "PHEV" else "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "7"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "28,75"
                    elif tipo_motor in ["PHEV", "HEV"]:
                        imesi_str = "2" if tipo_motor == "PHEV" else "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "7"

            # c) 1500 < cilindrada <= 2000
            elif 1500 < cilindrada <= 2000:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    elif tipo_motor == "PHEV":
                        imesi_str = "2"
                    elif tipo_motor == "HEV":
                        imesi_str = "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "14"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "34,5"
                    elif tipo_motor == "PHEV":
                        imesi_str = "2"
                    elif tipo_motor == "HEV":
                        imesi_str = "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "14"

            # d) 2000 < cilindrada <= 2500
            elif 2000 < cilindrada <= 2500:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    elif tipo_motor == "PHEV":
                        imesi_str = "2"
                    elif tipo_motor == "HEV":
                        imesi_str = "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "34,50"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "40,25"
                    elif tipo_motor == "PHEV":
                        imesi_str = "2"
                    elif tipo_motor == "HEV":
                        imesi_str = "3,45"
                    elif tipo_motor == "MHEV":
                        imesi_str = "34,50"

            # e) 2500 < cilindrada <= 3000
            elif 2500 < cilindrada <= 3000:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    else:
                        imesi_str = "34,50"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "40,25"
                    else:
                        imesi_str = "34,50"

            # f) cilindrada > 3000
            else:
                if combustible == "D":
                    if tipo_motor == "D":
                        imesi_str = "115"
                    else:
                        imesi_str = "34,50"
                elif combustible == "N":
                    if tipo_motor == "N":
                        imesi_str = "46"
                    else:
                        imesi_str = "34,50"

    # Si no entra en nada => 0%
    return porcentaje_str_a_decimal(imesi_str)


def porcentaje_str_a_decimal(porc_str):
    """
    Convierte un string que representa un porcentaje en formato con coma decimal
    o entero, a un número en formato fracción (0.xx).
    Ejemplos:
      "34,7"  -> 0.347
      "115"   -> 1.15
      "3,45"  -> 0.0345
      "6,0"   -> 0.06
      "2"     -> 0.02
    """
    porc_str = porc_str.replace(',', '.')
    try:
        val = float(porc_str)
    except:
        val = 0.0
    return val / 100.0

if __name__ == "__main__":
    file_path = r"C:\Users\emili\PycharmProjects\TesisUY\Base de datos 2023 - Prueba.xlsx"
    sheet_name = "2023"

    # Se llama a la función global del codigo
    procesar_imesi(file_path, sheet_name)